# -*- coding: utf-8 -*-

from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import openpyxl
from drive import *
import os
import shutil
import json
import sys
from openpyxl.worksheet.hyperlink import Hyperlink


SPREADSHEET_PATH = 'video_embed_spreadsheet.xlsx'
SPREADSHEET_ID = u'10p7bQcYlI5KEY93qduQTvR01ux63Qp6V'
SHEET_INDEX = 0
ID_COLUMN = 3
NAME_COLUMN = 10
VIDEO_COLUMN = 11
AUTO_LINK_MARK = u'[auto]'
ERROR_LINK_MARK = u'[error]'
NO_NAME_NAME = u'Video Link'
MIN_ROW = 4
MAX_ROW = 1287

REFERENCE_SPREADSHEET_PATH = 'reference_spreadsheet.xlsx'
REFERENCE_SHEET_INDEX = 0
REFERENCE_ID_COLUMN = 5
REFERENCE_VIDEO_COLUMN = 11

REFERENCE_SHEET_TO_INDEXES = {
	0: {'id': 5, 'video': 11, 'min_row': 3, 'max_row': 701, 'valid_links': True},
	1: {'id': 6, 'video': 12, 'min_row': 3, 'max_row': 574, 'valid_links': True},
	2: {'id': 1, 'video': 6, 'min_row': 2, 'max_row': 33, 'valid_links': False},
	3: {'id': 1, 'video': 7, 'min_row': 2, 'max_row': 55, 'valid_links': False}
}

TYPE_URL = 0
TYPE_ERROR_MSG = 1


class Db(object):

	def __init__(self, db_file):
		self._db_file = db_file
		
		if not os.path.exists(db_file):
			self._db_data = {
				'done': []
			}
			self.save()
		else:
			with open(db_file) as f:
				self._db_data = json.load(f)

	def save(self):
		with open(self._db_file, 'w') as a:
			json.dump(self._db_data, a)

	def update_id_done(self, id):
		self._db_data['done'].append(id)

	def was_id_done(self, id):
		return id in self._db_data['done']

def retrieve_spreadsheet(drive, out_path):
	local_file = out_path

	if os.path.exists(local_file):
		os.unlink(out_path)

	file = drive.CreateFile({'id': SPREADSHEET_ID})
	file.GetContentFile(out_path)

def stripped_id_value(raw_id):
	if type(raw_id) == unicode or type(raw_id) == str:
		if raw_id == '':
			raise ValueError('id is empty')

		raw_id = raw_id.strip()
		if raw_id.startswith('L') and not raw_id[1] == '-':
			if raw_id[1] == '_':
				raw_id = 'L-' + raw_id[2:]
			else:
				raw_id = 'L-' + raw_id[1:]
		if raw_id.startswith('M'):
			raw_id = raw_id[1:]

		return raw_id
	elif type(raw_id) == float:
		return str(int(raw_id))

	return str(raw_id)

def find_id_match_in_reference_sheet(id_value, reference_sheet, index_data):
	for i in xrange(index_data['min_row'], index_data['max_row']+1):
		try:
			raw_id = reference_sheet.cell(column=index_data['id'], row=i).value
			raw_id = stripped_id_value(raw_id)

			if raw_id == id_value:
				print 'Id match in reference', str(i)
				cell = reference_sheet.cell(column=index_data['video'], row=i)
				url = cell.hyperlink
				
				if url is None or url == '':
					url = cell.value
					if url is None or url == '':
						url = u'קישור חסר'
					return (TYPE_ERROR_MSG, url)

				if not index_data['valid_links']:
					return (TYPE_ERROR_MSG, u'לא הועלה (סיבה לא מצוינת)')

				if type(url) == Hyperlink:
					return (TYPE_URL, url.target)
				return (TYPE_URL, url)

		except Exception as e:
			print 'Error while checking for reference match', str(i), unicode(e)

	raise ValueError('Match not found for id ' + unicode(id_value))

def find_id_match_in_reference(id_value, reference_workbook):
	for i in xrange(len(reference_workbook.sheetnames)):
		if i >= len(REFERENCE_SHEET_TO_INDEXES):
			break

		try:
			sheetname = reference_workbook.sheetnames[i]
			print 'Checking in sheet', sheetname

			sheet = reference_workbook[sheetname]
			return find_id_match_in_reference_sheet(id_value, sheet, REFERENCE_SHEET_TO_INDEXES[i])
		except Exception as e:
			print 'Error with sheet check', unicode(e)

	raise ValueError('Match not found for id ' + str(id_value))


def update_spreadsheet(drive, refernce_file_path, out_path=SPREADSHEET_PATH, style='40 % - Accent1'):
	failed = []
	total = 0
	success = 0

	retrieve_spreadsheet(drive, out_path)

	reference_workbook = openpyxl.load_workbook(refernce_file_path)
	out_workbook = openpyxl.load_workbook(out_path)
	try:
		out_sheet = out_workbook[out_workbook.sheetnames[SHEET_INDEX]]

		for i in xrange(MIN_ROW, MAX_ROW+1):
			total += 1
			try:
				raw_id = out_sheet.cell(column=ID_COLUMN, row=i).value
				raw_id = stripped_id_value(raw_id)

				print 'Id:', raw_id, str(i)

				url_cell = out_sheet.cell(column=VIDEO_COLUMN, row=i)
				if url_cell.value not in ('', None):
					print 'Cell already has video link', str(i), '...skipping'
					success+=1
					continue

				val_type, video_link = find_id_match_in_reference(raw_id, reference_workbook)
				if val_type == TYPE_URL:
					name = out_sheet.cell(column=NAME_COLUMN, row=i).value
					if name is None or name.strip() == '':
						name = NO_NAME_NAME
					else:
						name = name.strip()

					name = AUTO_LINK_MARK + name

					print name
					print video_link

					url_cell.hyperlink = video_link
					url_cell.value = name
					url_cell.style = style
				else:
					name = ERROR_LINK_MARK + video_link
					url_cell.value = name
					url_cell.style = style

				success += 1
			except Exception as e:
				print 'Error handling row', str(i), unicode(e)

			print 'Done'
	finally:
		out_workbook.save(out_path)
		out_workbook.close()
		reference_workbook.close()

	print 'Amount ({0}/{1})'.format(success, total)


if not os.path.exists(REFERENCE_SPREADSHEET_PATH):
	print 'No reference spreadsheet'
	exit()

db = Db('video_embed.db.json')

gauth = GoogleAuth()
gauth.LocalWebserverAuth()
drive = GoogleDrive(gauth)

update_spreadsheet(drive, REFERENCE_SPREADSHEET_PATH)