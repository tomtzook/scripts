# -*- coding: utf-8 -*-

import openpyxl
from drive import *
import datetime
import os
#import gspread
#from oauth2client.service_account import ServiceAccountCredentials

SPREADSHEET_PATH = 'spreadsheet.xlsx'
SPREADSHEET_ID = u'10p7bQcYlI5KEY93qduQTvR01ux63Qp6V'
SHEET_INDEX = 0
DATE_COLUMN = 1
STUDIO_COLUMN = 2
ID_COLUMN = 3
URL_COLUMN = 12

def get_id_for_studio(studios, studio):
	return get_id_from_list_by_title(studios, studio)

def search_pdf_in_folder(drive, folder_id, fid):
	files = get_files_with_parentid(drive, folder_id)
	for f in files:
		if f['title'].find(fid) >= 0 and f['title'].endswith('.pdf'):
			return f
		if f['mimeType'] == u'application/vnd.google-apps.folder':
			result = search_pdf_in_folder(drive, f['id'], fid)
			if result is not None:
				return result

	return None


def get_properties_for_pdf(drive, studio_id, fid):
	result = search_pdf_in_folder(drive, studio_id, fid)
	if result is not None:
		return result
	raise ValueError('not found ' + fid)

def find_pdf_properties(drive, root_files, date, studio_name, fid):
	dateid = get_id_for_date(root_files, date)
	studios = get_files_with_parentid(drive, dateid)
	studioid = get_id_for_studio(studios, studio_name)
	return get_properties_for_pdf(drive, studioid, fid)

def retrieve_spreadsheet(drive, out_path):
	local_file = out_path

	if os.path.exists(local_file):
		os.unlink(out_path)

	file = drive.CreateFile({'id': SPREADSHEET_ID})
	file.GetContentFile(out_path)

def update_spreadsheet_for_date(drive, root_files, date_str, out_path=SPREADSHEET_PATH):
	found_start_of_date = False
	row_range = None
	failed = []

	retrieve_spreadsheet(drive, out_path)

	workbook = openpyxl.load_workbook(out_path)
	try:
		sheet = workbook[workbook.sheetnames[SHEET_INDEX]]
		for i in xrange(sheet.min_row, sheet.max_row+1):
			raw_date = sheet.cell(column=DATE_COLUMN, row=i).value
			if type(raw_date) != datetime.datetime:
				print 'date type error', i
				continue

			date = '{0}.{1}'.format(raw_date.day, raw_date.month)

			if date != date_str:
				if found_start_of_date:
					print 'Reached last row', i
					row_range = row_range + str(i)
					break
				continue

			if not found_start_of_date:
				print 'Reached first row', i
				row_range = str(i) + ' -> '
				found_start_of_date = True

			raw_studio = sheet.cell(column=STUDIO_COLUMN, row=i).value
			if raw_studio is None:
				print 'None studio'
				continue
				
			raw_studio = raw_studio.strip()

			raw_id = sheet.cell(column=ID_COLUMN, row=i).value

			if type(raw_id) == float:
				raw_id = int(raw_id)

			print date, raw_studio, raw_id, str(i)

			if raw_id == '':
				url_cell.value = ''
				url_cell.hyperlink = ''
				url_cell.style = 'Normal'
				continue

			try:
				props = find_pdf_properties(drive, root_files, date, raw_studio, str(raw_id))
				ff = drive.CreateFile({'id': props['id']})
				ff.FetchMetadata()

				print 'Shared:', ff.metadata.keys()
				print ff['capabilities']

				url_cell = sheet.cell(column=URL_COLUMN, row=i)
				url_cell.value = props['title'].split('.')[0]
				url_cell.hyperlink = props['alternateLink']
				url_cell.style = 'Normal'
			except Exception as e:
				failed.append((i, raw_id, e))
				try:
					print unicode(e)
				except:
					print 'error (cannot print specifics)'
				pass

		print 'Done update:', row_range, 'File:', out_path
		print 'Failed:\n\t', u'\n\t'.join([u'{0} - {1} ? {2}'.format(str(i), str(raw_id), unicode(e)) for i, raw_id, e in failed])
	finally:
		workbook.save(out_path)
		workbook.close()

def update_spreadsheet(drive, root_files, start_row, end_row):
	workbook = openpyxl.load_workbook(SPREADSHEET_PATH)
	try:
		sheet = workbook[workbook.sheetnames[SHEET_INDEX]]
		for i in xrange(start_row, end_row+1):
			raw_date = sheet.cell(column=DATE_COLUMN, row=i).value
			raw_studio = sheet.cell(column=STUDIO_COLUMN, row=i).value
			raw_id = sheet.cell(column=ID_COLUMN, row=i).value

			date = '{0}.{1}'.format(raw_date.day, raw_date.month)

			if type(raw_id) == float:
				raw_id = int(raw_id)

			print date, raw_studio, raw_id

			if raw_id == '':
				url_cell.value = ''
				url_cell.hyperlink = ''
				url_cell.style = 'Normal'
				continue

			try:
				props = find_pdf_properties(drive, root_files, date, raw_studio, str(raw_id))

				url_cell = sheet.cell(column=URL_COLUMN, row=i)
				url_cell.value = props['title'].split('.')[0]
				url_cell.hyperlink = props['alternateLink']
				url_cell.style = 'Normal'
			except Exception as e:
				try:
					print str(e)
				except:
					print 'error (cannot print specifics)'
				pass
	finally:
		workbook.save(SPREADSHEET_PATH)
		workbook.close()

#scope = ['https://spreadsheets.google.com/feeds']
#credentials = ServiceAccountCredentials.from_json_keyfile_name('client_secrets_service.json', scope)
#client = gspread.authorize(credentials)

#spreadsheet_url = 'https://docs.google.com/spreadsheets/d/10p7bQcYlI5KEY93qduQTvR01ux63Qp6V/edit'



#props = find_pdf_properties(drive, root_files, u'25.3', u'אולפן 4', u'477')
#print props